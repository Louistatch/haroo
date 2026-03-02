"""
Services pour le dashboard institutionnel
Exigences: 25.3, 25.4, 25.5, 25.6
"""
from decimal import Decimal
from django.db.models import Count, Sum, Avg, Q
from django.utils import timezone
from datetime import datetime, timedelta
import hashlib
import re
from apps.users.models import ExploitantProfile, AgronomeProfile, OuvrierProfile
from apps.payments.models import Transaction
from apps.locations.models import Region, Prefecture


class InstitutionalDashboardService:
    """
    Service pour générer les statistiques sectorielles du dashboard institutionnel
    """
    
    @staticmethod
    def get_aggregated_statistics(region_id=None, start_date=None, end_date=None):
        """
        Calcule les statistiques agrégées pour le dashboard institutionnel
        
        Args:
            region_id: ID de la région pour filtrer (optionnel)
            start_date: Date de début de la période (optionnel)
            end_date: Date de fin de la période (optionnel)
            
        Returns:
            dict: Statistiques agrégées
        """
        # Filtres de base
        exploitant_filter = Q(statut_verification='VERIFIE')
        agronome_filter = Q(statut_validation='VALIDE')
        transaction_filter = Q(statut='SUCCESS')
        
        # Filtre par région si spécifié
        if region_id:
            exploitant_filter &= Q(canton_principal__prefecture__region_id=region_id)
            agronome_filter &= Q(canton_rattachement__prefecture__region_id=region_id)
        
        # Filtre par période si spécifié
        if start_date:
            transaction_filter &= Q(created_at__gte=start_date)
        if end_date:
            transaction_filter &= Q(created_at__lte=end_date)
        
        # Nombre d'exploitations vérifiées
        nombre_exploitations = ExploitantProfile.objects.filter(
            exploitant_filter
        ).count()
        
        # Superficie totale cultivée
        superficie_totale = ExploitantProfile.objects.filter(
            exploitant_filter
        ).aggregate(
            total=Sum('superficie_totale')
        )['total'] or Decimal('0.00')
        
        # Emplois créés (agronomes validés + ouvriers)
        nombre_agronomes = AgronomeProfile.objects.filter(
            agronome_filter
        ).count()
        
        nombre_ouvriers = OuvrierProfile.objects.count()
        
        emplois_crees = nombre_agronomes + nombre_ouvriers
        
        # Volume et valeur des transactions
        transactions_stats = Transaction.objects.filter(
            transaction_filter
        ).aggregate(
            volume=Count('id'),
            valeur_totale=Sum('montant'),
            commission_totale=Sum('commission_plateforme')
        )
        
        volume_transactions = transactions_stats['volume'] or 0
        valeur_transactions = transactions_stats['valeur_totale'] or Decimal('0.00')
        commission_totale = transactions_stats['commission_totale'] or Decimal('0.00')
        
        return {
            'nombre_exploitations': nombre_exploitations,
            'superficie_totale_hectares': float(superficie_totale),
            'emplois_crees': {
                'total': emplois_crees,
                'agronomes': nombre_agronomes,
                'ouvriers': nombre_ouvriers
            },
            'transactions': {
                'volume': volume_transactions,
                'valeur_totale_fcfa': float(valeur_transactions),
                'commission_plateforme_fcfa': float(commission_totale)
            }
        }
    
    @staticmethod
    def get_statistics_by_region(start_date=None, end_date=None):
        """
        Calcule les statistiques par région
        
        Args:
            start_date: Date de début de la période (optionnel)
            end_date: Date de fin de la période (optionnel)
            
        Returns:
            list: Liste des statistiques par région
        """
        regions = Region.objects.all()
        statistics = []
        
        for region in regions:
            stats = InstitutionalDashboardService.get_aggregated_statistics(
                region_id=region.id,
                start_date=start_date,
                end_date=end_date
            )
            stats['region'] = {
                'id': region.id,
                'nom': region.nom,
                'code': region.code
            }
            statistics.append(stats)
        
        return statistics
    
    @staticmethod
    def get_statistics_by_prefecture(region_id=None, start_date=None, end_date=None):
        """
        Calcule les statistiques par préfecture
        
        Args:
            region_id: ID de la région pour filtrer (optionnel)
            start_date: Date de début de la période (optionnel)
            end_date: Date de fin de la période (optionnel)
            
        Returns:
            list: Liste des statistiques par préfecture
        """
        prefecture_filter = Q()
        if region_id:
            prefecture_filter = Q(region_id=region_id)
        
        prefectures = Prefecture.objects.filter(prefecture_filter)
        statistics = []
        
        for prefecture in prefectures:
            # Filtres pour cette préfecture
            exploitant_filter = Q(
                statut_verification='VERIFIE',
                canton_principal__prefecture_id=prefecture.id
            )
            agronome_filter = Q(
                statut_validation='VALIDE',
                canton_rattachement__prefecture_id=prefecture.id
            )
            
            # Filtre par période pour les transactions
            transaction_filter = Q(statut='SUCCESS')
            if start_date:
                transaction_filter &= Q(created_at__gte=start_date)
            if end_date:
                transaction_filter &= Q(created_at__lte=end_date)
            
            # Calculs
            nombre_exploitations = ExploitantProfile.objects.filter(
                exploitant_filter
            ).count()
            
            superficie_totale = ExploitantProfile.objects.filter(
                exploitant_filter
            ).aggregate(
                total=Sum('superficie_totale')
            )['total'] or Decimal('0.00')
            
            nombre_agronomes = AgronomeProfile.objects.filter(
                agronome_filter
            ).count()
            
            stats = {
                'prefecture': {
                    'id': prefecture.id,
                    'nom': prefecture.nom,
                    'code': prefecture.code,
                    'region': {
                        'id': prefecture.region.id,
                        'nom': prefecture.region.nom
                    }
                },
                'nombre_exploitations': nombre_exploitations,
                'superficie_totale_hectares': float(superficie_totale),
                'nombre_agronomes': nombre_agronomes
            }
            statistics.append(stats)
        
        return statistics
    
    @staticmethod
    def get_transaction_breakdown(region_id=None, start_date=None, end_date=None):
        """
        Obtient la répartition des transactions par type
        
        Args:
            region_id: ID de la région pour filtrer (optionnel)
            start_date: Date de début de la période (optionnel)
            end_date: Date de fin de la période (optionnel)
            
        Returns:
            list: Répartition des transactions par type
        """
        transaction_filter = Q(statut='SUCCESS')
        
        # Filtre par période
        if start_date:
            transaction_filter &= Q(created_at__gte=start_date)
        if end_date:
            transaction_filter &= Q(created_at__lte=end_date)
        
        # Répartition par type
        breakdown = Transaction.objects.filter(
            transaction_filter
        ).values('type_transaction').annotate(
            nombre=Count('id'),
            montant_total=Sum('montant'),
            commission_totale=Sum('commission_plateforme')
        ).order_by('-montant_total')
        
        return [
            {
                'type': item['type_transaction'],
                'nombre_transactions': item['nombre'],
                'montant_total_fcfa': float(item['montant_total'] or 0),
                'commission_totale_fcfa': float(item['commission_totale'] or 0)
            }
            for item in breakdown
        ]
    
    @staticmethod
    def get_monthly_trends(region_id=None, months=12):
        """
        Obtient les tendances mensuelles
        
        Args:
            region_id: ID de la région pour filtrer (optionnel)
            months: Nombre de mois à inclure (défaut: 12)
            
        Returns:
            list: Tendances mensuelles
        """
        end_date = timezone.now()
        start_date = end_date - timedelta(days=months * 30)
        
        # Générer les statistiques pour chaque mois
        trends = []
        current_date = start_date
        
        while current_date <= end_date:
            month_start = current_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            # Dernier jour du mois
            if month_start.month == 12:
                month_end = month_start.replace(year=month_start.year + 1, month=1, day=1) - timedelta(seconds=1)
            else:
                month_end = month_start.replace(month=month_start.month + 1, day=1) - timedelta(seconds=1)
            
            stats = InstitutionalDashboardService.get_aggregated_statistics(
                region_id=region_id,
                start_date=month_start,
                end_date=month_end
            )
            
            stats['periode'] = {
                'annee': month_start.year,
                'mois': month_start.month,
                'mois_nom': month_start.strftime('%B'),
                'debut': month_start.isoformat(),
                'fin': month_end.isoformat()
            }
            
            trends.append(stats)
            
            # Passer au mois suivant
            if current_date.month == 12:
                current_date = current_date.replace(year=current_date.year + 1, month=1)
            else:
                current_date = current_date.replace(month=current_date.month + 1)
        
        return trends



class DataAnonymizationService:
    """
    Service pour anonymiser les données personnelles dans les exports institutionnels
    Exigences: 25.6
    """
    
    @staticmethod
    def anonymize_phone_number(phone_number):
        """
        Anonymise un numéro de téléphone en masquant les chiffres du milieu
        
        Args:
            phone_number (str): Numéro de téléphone à anonymiser
            
        Returns:
            str: Numéro de téléphone anonymisé (ex: +228 XX XX XX 89)
        """
        if not phone_number:
            return "N/A"
        
        # Garder les 3 premiers et 2 derniers chiffres
        phone_str = str(phone_number)
        if len(phone_str) <= 5:
            return "XXX"
        
        # Format: +228 XX XX XX 89
        prefix = phone_str[:4] if phone_str.startswith('+') else phone_str[:3]
        suffix = phone_str[-2:]
        masked_middle = " XX XX XX "
        
        return f"{prefix}{masked_middle}{suffix}"
    
    @staticmethod
    def anonymize_email(email):
        """
        Anonymise une adresse email en masquant une partie
        
        Args:
            email (str): Email à anonymiser
            
        Returns:
            str: Email anonymisé (ex: j***@example.com)
        """
        if not email:
            return "N/A"
        
        try:
            local, domain = email.split('@')
            if len(local) <= 1:
                return f"*@{domain}"
            return f"{local[0]}***@{domain}"
        except ValueError:
            return "***@***.***"
    
    @staticmethod
    def anonymize_name(name):
        """
        Anonymise un nom en ne gardant que l'initiale
        
        Args:
            name (str): Nom à anonymiser
            
        Returns:
            str: Nom anonymisé (ex: "Jean Dupont" -> "J. D.")
        """
        if not name:
            return "N/A"
        
        parts = name.strip().split()
        if not parts:
            return "N/A"
        
        # Garder uniquement les initiales
        initials = [part[0].upper() + "." for part in parts if part]
        return " ".join(initials)
    
    @staticmethod
    def anonymize_username(username):
        """
        Anonymise un nom d'utilisateur en utilisant un hash
        
        Args:
            username (str): Nom d'utilisateur à anonymiser
            
        Returns:
            str: Identifiant anonymisé (ex: "USER_a3f5b2")
        """
        if not username:
            return "USER_ANONYME"
        
        # Créer un hash court du nom d'utilisateur
        hash_obj = hashlib.md5(username.encode())
        hash_short = hash_obj.hexdigest()[:6]
        return f"USER_{hash_short}"
    
    @staticmethod
    def anonymize_gps_coordinates(coordinates, precision=2):
        """
        Réduit la précision des coordonnées GPS pour anonymiser la localisation exacte
        
        Args:
            coordinates: Coordonnées GPS (dict avec lat/lon ou objet Point)
            precision (int): Nombre de décimales à conserver (défaut: 2)
            
        Returns:
            dict: Coordonnées GPS avec précision réduite
        """
        if not coordinates:
            return {"lat": None, "lon": None}
        
        try:
            # Si c'est un objet Point PostGIS
            if hasattr(coordinates, 'x') and hasattr(coordinates, 'y'):
                lat = round(coordinates.y, precision)
                lon = round(coordinates.x, precision)
            # Si c'est un dict
            elif isinstance(coordinates, dict):
                lat = round(float(coordinates.get('lat', 0)), precision)
                lon = round(float(coordinates.get('lon', 0)), precision)
            else:
                return {"lat": None, "lon": None}
            
            return {"lat": lat, "lon": lon}
        except (ValueError, TypeError, AttributeError):
            return {"lat": None, "lon": None}
    
    @staticmethod
    def anonymize_user_data(user_data):
        """
        Anonymise toutes les données personnelles d'un utilisateur
        
        Args:
            user_data (dict): Données utilisateur à anonymiser
            
        Returns:
            dict: Données utilisateur anonymisées
        """
        anonymized = user_data.copy()
        
        # Anonymiser les champs personnels
        if 'username' in anonymized:
            anonymized['username'] = DataAnonymizationService.anonymize_username(
                anonymized['username']
            )
        
        if 'first_name' in anonymized:
            anonymized['first_name'] = DataAnonymizationService.anonymize_name(
                anonymized['first_name']
            )
        
        if 'last_name' in anonymized:
            anonymized['last_name'] = DataAnonymizationService.anonymize_name(
                anonymized['last_name']
            )
        
        if 'email' in anonymized:
            anonymized['email'] = DataAnonymizationService.anonymize_email(
                anonymized['email']
            )
        
        if 'phone_number' in anonymized:
            anonymized['phone_number'] = DataAnonymizationService.anonymize_phone_number(
                anonymized['phone_number']
            )
        
        if 'coordonnees_gps' in anonymized:
            anonymized['coordonnees_gps'] = DataAnonymizationService.anonymize_gps_coordinates(
                anonymized['coordonnees_gps']
            )
        
        # Supprimer les champs sensibles
        sensitive_fields = [
            'password', 'two_factor_secret', 'photo_profil',
            'date_joined', 'last_login', 'id'
        ]
        for field in sensitive_fields:
            anonymized.pop(field, None)
        
        return anonymized
    
    @staticmethod
    def anonymize_statistics_export(statistics_data):
        """
        Anonymise les données statistiques pour l'export
        
        Args:
            statistics_data (dict): Données statistiques à anonymiser
            
        Returns:
            dict: Données statistiques anonymisées
        """
        anonymized = {}
        
        # Copier les données agrégées (pas de données personnelles)
        for key, value in statistics_data.items():
            if isinstance(value, dict):
                anonymized[key] = DataAnonymizationService.anonymize_statistics_export(value)
            elif isinstance(value, list):
                anonymized[key] = [
                    DataAnonymizationService.anonymize_statistics_export(item)
                    if isinstance(item, dict) else item
                    for item in value
                ]
            else:
                # Les valeurs numériques et les noms de régions/préfectures sont OK
                anonymized[key] = value
        
        return anonymized
    
    @staticmethod
    def prepare_export_data(statistics, include_details=False):
        """
        Prépare les données pour l'export en les anonymisant
        
        Args:
            statistics (dict): Statistiques à exporter
            include_details (bool): Inclure les détails par région/préfecture
            
        Returns:
            dict: Données prêtes pour l'export
        """
        export_data = {
            'date_export': timezone.now().isoformat(),
            'type_export': 'Statistiques Sectorielles Anonymisées',
            'avertissement': 'Toutes les données personnelles ont été anonymisées conformément à la réglementation',
            'statistiques': DataAnonymizationService.anonymize_statistics_export(statistics)
        }
        
        return export_data



class ReportGenerationService:
    """
    Service pour générer des rapports Excel et PDF anonymisés
    Exigences: 25.5, 25.6
    """
    
    @staticmethod
    def generate_excel_report(statistics_data, filename='rapport_statistiques.xlsx'):
        """
        Génère un rapport Excel avec les statistiques anonymisées
        
        Args:
            statistics_data (dict): Données statistiques anonymisées
            filename (str): Nom du fichier Excel
            
        Returns:
            BytesIO: Fichier Excel en mémoire
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO
        
        # Créer un nouveau workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Statistiques Sectorielles"
        
        # Style pour les en-têtes
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Titre du rapport
        ws['A1'] = "RAPPORT STATISTIQUES SECTORIELLES - PLATEFORME AGRICOLE TOGO"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        # Date d'export
        ws['A2'] = f"Date d'export: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        ws.merge_cells('A2:D2')
        
        # Avertissement d'anonymisation
        ws['A3'] = "⚠️ Toutes les données personnelles ont été anonymisées"
        ws['A3'].font = Font(italic=True, color="FF0000")
        ws.merge_cells('A3:D3')
        
        row = 5
        
        # Section: Statistiques Globales
        if 'statistiques_globales' in statistics_data or 'nombre_exploitations' in statistics_data:
            stats = statistics_data.get('statistiques_globales', statistics_data)
            
            ws[f'A{row}'] = "STATISTIQUES GLOBALES"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            # Nombre d'exploitations
            ws[f'A{row}'] = "Nombre d'exploitations vérifiées"
            ws[f'B{row}'] = stats.get('nombre_exploitations', 0)
            row += 1
            
            # Superficie totale
            ws[f'A{row}'] = "Superficie totale cultivée (hectares)"
            ws[f'B{row}'] = stats.get('superficie_totale_hectares', 0)
            row += 1
            
            # Emplois créés
            if 'emplois_crees' in stats:
                emplois = stats['emplois_crees']
                ws[f'A{row}'] = "Emplois créés (total)"
                ws[f'B{row}'] = emplois.get('total', 0)
                row += 1
                
                ws[f'A{row}'] = "  - Agronomes validés"
                ws[f'B{row}'] = emplois.get('agronomes', 0)
                row += 1
                
                ws[f'A{row}'] = "  - Ouvriers agricoles"
                ws[f'B{row}'] = emplois.get('ouvriers', 0)
                row += 1
            
            # Transactions
            if 'transactions' in stats:
                transactions = stats['transactions']
                ws[f'A{row}'] = "Volume de transactions"
                ws[f'B{row}'] = transactions.get('volume', 0)
                row += 1
                
                ws[f'A{row}'] = "Valeur totale (FCFA)"
                ws[f'B{row}'] = transactions.get('valeur_totale_fcfa', 0)
                row += 1
                
                ws[f'A{row}'] = "Commission plateforme (FCFA)"
                ws[f'B{row}'] = transactions.get('commission_plateforme_fcfa', 0)
                row += 1
            
            row += 2
        
        # Section: Statistiques par Région
        if 'statistiques_par_region' in statistics_data:
            ws[f'A{row}'] = "STATISTIQUES PAR RÉGION"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
            
            # En-têtes du tableau
            headers = ['Région', 'Exploitations', 'Superficie (ha)', 'Emplois', 'Transactions']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = header_alignment
            row += 1
            
            # Données par région
            for region_stats in statistics_data['statistiques_par_region']:
                region_info = region_stats.get('region', {})
                emplois = region_stats.get('emplois_crees', {})
                transactions = region_stats.get('transactions', {})
                
                ws[f'A{row}'] = region_info.get('nom', 'N/A')
                ws[f'B{row}'] = region_stats.get('nombre_exploitations', 0)
                ws[f'C{row}'] = region_stats.get('superficie_totale_hectares', 0)
                ws[f'D{row}'] = emplois.get('total', 0)
                ws[f'E{row}'] = transactions.get('volume', 0)
                row += 1
            
            row += 2
        
        # Section: Répartition des Transactions
        if 'repartition_transactions' in statistics_data:
            ws[f'A{row}'] = "RÉPARTITION DES TRANSACTIONS PAR TYPE"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            
            # En-têtes
            headers = ['Type', 'Nombre', 'Montant Total (FCFA)', 'Commission (FCFA)']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = header_alignment
            row += 1
            
            # Données
            for transaction_type in statistics_data['repartition_transactions']:
                ws[f'A{row}'] = transaction_type.get('type', 'N/A')
                ws[f'B{row}'] = transaction_type.get('nombre_transactions', 0)
                ws[f'C{row}'] = transaction_type.get('montant_total_fcfa', 0)
                ws[f'D{row}'] = transaction_type.get('commission_totale_fcfa', 0)
                row += 1
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        
        # Sauvegarder dans un BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def generate_pdf_report(statistics_data, filename='rapport_statistiques.pdf'):
        """
        Génère un rapport PDF avec les statistiques anonymisées
        
        Args:
            statistics_data (dict): Données statistiques anonymisées
            filename (str): Nom du fichier PDF
            
        Returns:
            BytesIO: Fichier PDF en mémoire
        """
        from reportlab.lib.pagesizes import A4, letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from io import BytesIO
        
        # Créer un buffer en mémoire
        buffer = BytesIO()
        
        # Créer le document PDF
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#366092'),
            spaceAfter=12,
            spaceBefore=12
        )
        
        warning_style = ParagraphStyle(
            'Warning',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.red,
            alignment=TA_CENTER,
            spaceAfter=20
        )
        
        # Titre
        title = Paragraph(
            "RAPPORT STATISTIQUES SECTORIELLES<br/>PLATEFORME AGRICOLE TOGO",
            title_style
        )
        story.append(title)
        
        # Date d'export
        date_text = Paragraph(
            f"Date d'export: {timezone.now().strftime('%d/%m/%Y %H:%M')}",
            styles['Normal']
        )
        story.append(date_text)
        story.append(Spacer(1, 0.2*inch))
        
        # Avertissement
        warning = Paragraph(
            "⚠️ Toutes les données personnelles ont été anonymisées conformément à la réglementation",
            warning_style
        )
        story.append(warning)
        story.append(Spacer(1, 0.3*inch))
        
        # Section: Statistiques Globales
        if 'statistiques_globales' in statistics_data or 'nombre_exploitations' in statistics_data:
            stats = statistics_data.get('statistiques_globales', statistics_data)
            
            heading = Paragraph("STATISTIQUES GLOBALES", heading_style)
            story.append(heading)
            
            # Créer un tableau pour les statistiques
            data = [
                ['Indicateur', 'Valeur']
            ]
            
            data.append(['Nombre d\'exploitations vérifiées', str(stats.get('nombre_exploitations', 0))])
            data.append(['Superficie totale cultivée (ha)', f"{stats.get('superficie_totale_hectares', 0):.2f}"])
            
            if 'emplois_crees' in stats:
                emplois = stats['emplois_crees']
                data.append(['Emplois créés (total)', str(emplois.get('total', 0))])
                data.append(['  - Agronomes validés', str(emplois.get('agronomes', 0))])
                data.append(['  - Ouvriers agricoles', str(emplois.get('ouvriers', 0))])
            
            if 'transactions' in stats:
                transactions = stats['transactions']
                data.append(['Volume de transactions', str(transactions.get('volume', 0))])
                data.append(['Valeur totale (FCFA)', f"{transactions.get('valeur_totale_fcfa', 0):,.0f}"])
                data.append(['Commission plateforme (FCFA)', f"{transactions.get('commission_plateforme_fcfa', 0):,.0f}"])
            
            table = Table(data, colWidths=[4*inch, 2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            story.append(Spacer(1, 0.3*inch))
        
        # Section: Statistiques par Région
        if 'statistiques_par_region' in statistics_data:
            heading = Paragraph("STATISTIQUES PAR RÉGION", heading_style)
            story.append(heading)
            
            data = [
                ['Région', 'Exploitations', 'Superficie (ha)', 'Emplois', 'Transactions']
            ]
            
            for region_stats in statistics_data['statistiques_par_region']:
                region_info = region_stats.get('region', {})
                emplois = region_stats.get('emplois_crees', {})
                transactions = region_stats.get('transactions', {})
                
                data.append([
                    region_info.get('nom', 'N/A'),
                    str(region_stats.get('nombre_exploitations', 0)),
                    f"{region_stats.get('superficie_totale_hectares', 0):.2f}",
                    str(emplois.get('total', 0)),
                    str(transactions.get('volume', 0))
                ])
            
            table = Table(data, colWidths=[1.5*inch, 1.2*inch, 1.2*inch, 1*inch, 1.2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            story.append(Spacer(1, 0.3*inch))
        
        # Section: Répartition des Transactions
        if 'repartition_transactions' in statistics_data:
            heading = Paragraph("RÉPARTITION DES TRANSACTIONS PAR TYPE", heading_style)
            story.append(heading)
            
            data = [
                ['Type', 'Nombre', 'Montant Total (FCFA)', 'Commission (FCFA)']
            ]
            
            for transaction_type in statistics_data['repartition_transactions']:
                data.append([
                    transaction_type.get('type', 'N/A'),
                    str(transaction_type.get('nombre_transactions', 0)),
                    f"{transaction_type.get('montant_total_fcfa', 0):,.0f}",
                    f"{transaction_type.get('commission_totale_fcfa', 0):,.0f}"
                ])
            
            table = Table(data, colWidths=[2*inch, 1.2*inch, 1.8*inch, 1.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
        
        # Construire le PDF
        doc.build(story)
        
        # Récupérer le contenu du buffer
        buffer.seek(0)
        return buffer
