"""
Tests pour le dashboard institutionnel
Exigences: 25.1, 25.2, 25.3, 25.4
"""
from decimal import Decimal
from django.test import TestCase
from django.contrib.auth import get_user_model
from django.utils import timezone
from datetime import timedelta
from rest_framework.test import APIClient
from rest_framework import status

from apps.locations.models import Region, Prefecture, Canton
from apps.users.models import ExploitantProfile, AgronomeProfile, OuvrierProfile, InstitutionProfile
from apps.payments.models import Transaction
from .services import InstitutionalDashboardService

User = get_user_model()


class InstitutionalDashboardServiceTests(TestCase):
    """Tests pour InstitutionalDashboardService"""
    
    def setUp(self):
        """Configuration des données de test"""
        # Créer une région, préfecture et canton
        self.region = Region.objects.create(nom="Maritime", code="MAR")
        self.prefecture = Prefecture.objects.create(
            nom="Golfe",
            code="GOL",
            region=self.region
        )
        self.canton = Canton.objects.create(
            nom="Lomé 1er",
            code="LOM1",
            prefecture=self.prefecture
        )
        
        # Créer des utilisateurs exploitants
        self.exploitant_user = User.objects.create_user(
            username='exploitant1',
            phone_number='+22890000001',
            user_type='EXPLOITANT',
            password='Test1234!'
        )
        self.exploitant_profile = ExploitantProfile.objects.create(
            user=self.exploitant_user,
            superficie_totale=Decimal('15.50'),
            canton_principal=self.canton,
            coordonnees_gps={'lat': 6.1319, 'lon': 1.2228},
            statut_verification='VERIFIE',
            cultures_actuelles=['Maïs', 'Tomate']
        )
        
        # Créer un agronome
        self.agronome_user = User.objects.create_user(
            username='agronome1',
            phone_number='+22890000002',
            user_type='AGRONOME',
            password='Test1234!'
        )
        self.agronome_profile = AgronomeProfile.objects.create(
            user=self.agronome_user,
            canton_rattachement=self.canton,
            specialisations=['Maraîchage', 'Irrigation'],
            statut_validation='VALIDE',
            badge_valide=True
        )
        
        # Créer un ouvrier
        self.ouvrier_user = User.objects.create_user(
            username='ouvrier1',
            phone_number='+22890000003',
            user_type='OUVRIER',
            password='Test1234!'
        )
        self.ouvrier_profile = OuvrierProfile.objects.create(
            user=self.ouvrier_user,
            competences=['Récolte', 'Plantation']
        )
        
        # Créer des transactions
        self.transaction1 = Transaction.objects.create(
            utilisateur=self.exploitant_user,
            type_transaction='ACHAT_DOCUMENT',
            montant=Decimal('5000.00'),
            commission_plateforme=Decimal('0.00'),
            statut='SUCCESS'
        )
        
        self.transaction2 = Transaction.objects.create(
            utilisateur=self.exploitant_user,
            type_transaction='RECRUTEMENT_AGRONOME',
            montant=Decimal('50000.00'),
            commission_plateforme=Decimal('5000.00'),
            statut='SUCCESS'
        )
    
    def test_get_aggregated_statistics(self):
        """Test du calcul des statistiques agrégées"""
        stats = InstitutionalDashboardService.get_aggregated_statistics()
        
        self.assertEqual(stats['nombre_exploitations'], 1)
        self.assertEqual(stats['superficie_totale_hectares'], 15.50)
        self.assertEqual(stats['emplois_crees']['total'], 2)
        self.assertEqual(stats['emplois_crees']['agronomes'], 1)
        self.assertEqual(stats['emplois_crees']['ouvriers'], 1)
        self.assertEqual(stats['transactions']['volume'], 2)
        self.assertEqual(stats['transactions']['valeur_totale_fcfa'], 55000.00)
        self.assertEqual(stats['transactions']['commission_plateforme_fcfa'], 5000.00)
    
    def test_get_aggregated_statistics_with_region_filter(self):
        """Test du filtrage par région"""
        stats = InstitutionalDashboardService.get_aggregated_statistics(
            region_id=self.region.id
        )
        
        self.assertEqual(stats['nombre_exploitations'], 1)
        self.assertEqual(stats['superficie_totale_hectares'], 15.50)
    
    def test_get_aggregated_statistics_with_date_filter(self):
        """Test du filtrage par période"""
        # Transactions créées aujourd'hui
        today = timezone.now()
        yesterday = today - timedelta(days=1)
        tomorrow = today + timedelta(days=1)
        
        stats = InstitutionalDashboardService.get_aggregated_statistics(
            start_date=yesterday,
            end_date=tomorrow
        )
        
        self.assertEqual(stats['transactions']['volume'], 2)
    
    def test_get_statistics_by_region(self):
        """Test des statistiques par région"""
        stats_list = InstitutionalDashboardService.get_statistics_by_region()
        
        self.assertIsInstance(stats_list, list)
        self.assertGreater(len(stats_list), 0)
        
        # Vérifier la structure
        first_stat = stats_list[0]
        self.assertIn('region', first_stat)
        self.assertIn('nombre_exploitations', first_stat)
        self.assertIn('superficie_totale_hectares', first_stat)
    
    def test_get_statistics_by_prefecture(self):
        """Test des statistiques par préfecture"""
        stats_list = InstitutionalDashboardService.get_statistics_by_prefecture(
            region_id=self.region.id
        )
        
        self.assertIsInstance(stats_list, list)
        self.assertGreater(len(stats_list), 0)
        
        # Vérifier la structure
        first_stat = stats_list[0]
        self.assertIn('prefecture', first_stat)
        self.assertIn('nombre_exploitations', first_stat)
        self.assertIn('superficie_totale_hectares', first_stat)
        self.assertIn('nombre_agronomes', first_stat)
    
    def test_get_transaction_breakdown(self):
        """Test de la répartition des transactions"""
        breakdown = InstitutionalDashboardService.get_transaction_breakdown()
        
        self.assertIsInstance(breakdown, list)
        self.assertEqual(len(breakdown), 2)  # 2 types de transactions
        
        # Vérifier la structure
        for item in breakdown:
            self.assertIn('type', item)
            self.assertIn('nombre_transactions', item)
            self.assertIn('montant_total_fcfa', item)
            self.assertIn('commission_totale_fcfa', item)


class InstitutionalDashboardAPITests(TestCase):
    """Tests pour les endpoints API du dashboard institutionnel"""
    
    def setUp(self):
        """Configuration des données de test"""
        self.client = APIClient()
        
        # Créer une région
        self.region = Region.objects.create(nom="Maritime", code="MAR")
        
        # Créer un utilisateur institutionnel avec 2FA
        self.institution_user = User.objects.create_user(
            username='institution1',
            phone_number='+22890000010',
            user_type='INSTITUTION',
            password='Test1234!',
            two_factor_enabled=True
        )
        self.institution_profile = InstitutionProfile.objects.create(
            user=self.institution_user,
            nom_organisme='Ministère de l\'Agriculture',
            niveau_acces='NATIONAL'
        )
        
        # Créer un utilisateur non-institutionnel
        self.regular_user = User.objects.create_user(
            username='regular1',
            phone_number='+22890000011',
            user_type='EXPLOITANT',
            password='Test1234!'
        )
    
    def test_dashboard_requires_authentication(self):
        """Test que le dashboard nécessite une authentification"""
        response = self.client.get('/api/v1/institutional/dashboard/')
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)
    
    def test_dashboard_requires_institutional_user(self):
        """Test que le dashboard nécessite un compte institutionnel"""
        self.client.force_authenticate(user=self.regular_user)
        response = self.client.get('/api/v1/institutional/dashboard/')
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
    
    def test_dashboard_requires_2fa(self):
        """Test que le dashboard nécessite le 2FA activé"""
        # Désactiver le 2FA
        self.institution_user.two_factor_enabled = False
        self.institution_user.save()
        
        self.client.force_authenticate(user=self.institution_user)
        response = self.client.get('/api/v1/institutional/dashboard/')
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
    
    def test_dashboard_success(self):
        """Test d'accès réussi au dashboard"""
        self.client.force_authenticate(user=self.institution_user)
        response = self.client.get('/api/v1/institutional/dashboard/')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('statistiques_globales', response.data)
        self.assertIn('statistiques_par_region', response.data)
        self.assertIn('repartition_transactions', response.data)
        self.assertIn('filtres_appliques', response.data)
    
    def test_dashboard_with_region_filter(self):
        """Test du dashboard avec filtre par région"""
        self.client.force_authenticate(user=self.institution_user)
        response = self.client.get(
            f'/api/v1/institutional/dashboard/?region={self.region.id}'
        )
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response.data['filtres_appliques']['region_id'],
            self.region.id
        )
    
    def test_aggregated_statistics_endpoint(self):
        """Test de l'endpoint des statistiques agrégées"""
        self.client.force_authenticate(user=self.institution_user)
        response = self.client.get('/api/v1/institutional/statistics/aggregated/')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('nombre_exploitations', response.data)
        self.assertIn('superficie_totale_hectares', response.data)
        self.assertIn('emplois_crees', response.data)
        self.assertIn('transactions', response.data)
    
    def test_statistics_by_prefecture_endpoint(self):
        """Test de l'endpoint des statistiques par préfecture"""
        self.client.force_authenticate(user=self.institution_user)
        response = self.client.get('/api/v1/institutional/statistics/by-prefecture/')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIsInstance(response.data, list)
    
    def test_transaction_breakdown_endpoint(self):
        """Test de l'endpoint de répartition des transactions"""
        self.client.force_authenticate(user=self.institution_user)
        response = self.client.get('/api/v1/institutional/statistics/transactions/')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIsInstance(response.data, list)
    
    def test_monthly_trends_endpoint(self):
        """Test de l'endpoint des tendances mensuelles"""
        self.client.force_authenticate(user=self.institution_user)
        response = self.client.get('/api/v1/institutional/statistics/trends/')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIsInstance(response.data, list)



class DataAnonymizationServiceTests(TestCase):
    """
    Tests pour le service d'anonymisation des données
    Exigences: 25.6
    """
    
    def setUp(self):
        """Configuration des données de test"""
        from .services import DataAnonymizationService
        self.service = DataAnonymizationService
    
    def test_anonymize_phone_number(self):
        """Test de l'anonymisation des numéros de téléphone"""
        # Numéro togolais standard
        phone = "+22890123456"
        anonymized = self.service.anonymize_phone_number(phone)
        
        # Vérifier que le numéro est masqué
        self.assertIn("XX", anonymized)
        self.assertTrue(anonymized.startswith("+228"))
        self.assertTrue(anonymized.endswith("56"))
        
        # Numéro vide
        self.assertEqual(self.service.anonymize_phone_number(""), "N/A")
        self.assertEqual(self.service.anonymize_phone_number(None), "N/A")
    
    def test_anonymize_email(self):
        """Test de l'anonymisation des emails"""
        # Email standard
        email = "jean.dupont@example.com"
        anonymized = self.service.anonymize_email(email)
        
        # Vérifier que l'email est masqué
        self.assertTrue(anonymized.startswith("j***"))
        self.assertIn("@example.com", anonymized)
        
        # Email court
        short_email = "a@test.com"
        anonymized_short = self.service.anonymize_email(short_email)
        self.assertTrue(anonymized_short.startswith("*@"))
        
        # Email vide
        self.assertEqual(self.service.anonymize_email(""), "N/A")
        self.assertEqual(self.service.anonymize_email(None), "N/A")
    
    def test_anonymize_name(self):
        """Test de l'anonymisation des noms"""
        # Nom complet
        name = "Jean Dupont"
        anonymized = self.service.anonymize_name(name)
        self.assertEqual(anonymized, "J. D.")
        
        # Nom simple
        simple_name = "Jean"
        anonymized_simple = self.service.anonymize_name(simple_name)
        self.assertEqual(anonymized_simple, "J.")
        
        # Nom avec plusieurs parties
        long_name = "Jean Pierre Dupont Martin"
        anonymized_long = self.service.anonymize_name(long_name)
        self.assertEqual(anonymized_long, "J. P. D. M.")
        
        # Nom vide
        self.assertEqual(self.service.anonymize_name(""), "N/A")
        self.assertEqual(self.service.anonymize_name(None), "N/A")
    
    def test_anonymize_username(self):
        """Test de l'anonymisation des noms d'utilisateur"""
        username = "exploitant123"
        anonymized = self.service.anonymize_username(username)
        
        # Vérifier le format
        self.assertTrue(anonymized.startswith("USER_"))
        self.assertEqual(len(anonymized), 11)  # USER_ + 6 caractères de hash
        
        # Vérifier la cohérence (même username = même hash)
        anonymized2 = self.service.anonymize_username(username)
        self.assertEqual(anonymized, anonymized2)
        
        # Username différent = hash différent
        different = self.service.anonymize_username("autre_user")
        self.assertNotEqual(anonymized, different)
        
        # Username vide
        self.assertEqual(self.service.anonymize_username(""), "USER_ANONYME")
        self.assertEqual(self.service.anonymize_username(None), "USER_ANONYME")
    
    def test_anonymize_gps_coordinates_dict(self):
        """Test de l'anonymisation des coordonnées GPS (format dict)"""
        # Coordonnées précises
        coords = {"lat": 6.131944, "lon": 1.222778}
        anonymized = self.service.anonymize_gps_coordinates(coords, precision=2)
        
        # Vérifier la réduction de précision
        self.assertEqual(anonymized["lat"], 6.13)
        self.assertEqual(anonymized["lon"], 1.22)
        
        # Coordonnées vides
        self.assertEqual(
            self.service.anonymize_gps_coordinates(None),
            {"lat": None, "lon": None}
        )
        self.assertEqual(
            self.service.anonymize_gps_coordinates({}),
            {"lat": None, "lon": None}
        )
    
    def test_anonymize_user_data(self):
        """Test de l'anonymisation complète des données utilisateur"""
        user_data = {
            'id': 123,
            'username': 'exploitant123',
            'first_name': 'Jean',
            'last_name': 'Dupont',
            'email': 'jean.dupont@example.com',
            'phone_number': '+22890123456',
            'coordonnees_gps': {'lat': 6.131944, 'lon': 1.222778},
            'password': 'hashed_password',
            'two_factor_secret': 'secret_key',
            'photo_profil': '/media/profiles/photo.jpg',
            'user_type': 'EXPLOITANT'
        }
        
        anonymized = self.service.anonymize_user_data(user_data)
        
        # Vérifier que les données personnelles sont anonymisées
        self.assertTrue(anonymized['username'].startswith('USER_'))
        self.assertEqual(anonymized['first_name'], 'J.')
        self.assertEqual(anonymized['last_name'], 'D.')
        self.assertIn('***', anonymized['email'])
        self.assertIn('XX', anonymized['phone_number'])
        self.assertEqual(anonymized['coordonnees_gps']['lat'], 6.13)
        
        # Vérifier que les champs sensibles sont supprimés
        self.assertNotIn('id', anonymized)
        self.assertNotIn('password', anonymized)
        self.assertNotIn('two_factor_secret', anonymized)
        self.assertNotIn('photo_profil', anonymized)
        
        # Vérifier que les champs non-sensibles sont conservés
        self.assertEqual(anonymized['user_type'], 'EXPLOITANT')
    
    def test_anonymize_statistics_export(self):
        """Test de l'anonymisation des données statistiques"""
        statistics = {
            'nombre_exploitations': 150,
            'superficie_totale_hectares': 2500.50,
            'emplois_crees': {
                'total': 200,
                'agronomes': 50,
                'ouvriers': 150
            },
            'transactions': {
                'volume': 500,
                'valeur_totale_fcfa': 50000000.00
            },
            'region': {
                'id': 1,
                'nom': 'Maritime',
                'code': 'MAR'
            }
        }
        
        anonymized = self.service.anonymize_statistics_export(statistics)
        
        # Vérifier que les données agrégées sont conservées
        self.assertEqual(anonymized['nombre_exploitations'], 150)
        self.assertEqual(anonymized['superficie_totale_hectares'], 2500.50)
        self.assertEqual(anonymized['emplois_crees']['total'], 200)
        self.assertEqual(anonymized['transactions']['volume'], 500)
        
        # Vérifier que les informations de région sont conservées (pas de données personnelles)
        self.assertEqual(anonymized['region']['nom'], 'Maritime')
    
    def test_prepare_export_data(self):
        """Test de la préparation des données pour l'export"""
        statistics = {
            'nombre_exploitations': 100,
            'superficie_totale_hectares': 1500.00
        }
        
        export_data = self.service.prepare_export_data(statistics)
        
        # Vérifier la structure
        self.assertIn('date_export', export_data)
        self.assertIn('type_export', export_data)
        self.assertIn('avertissement', export_data)
        self.assertIn('statistiques', export_data)
        
        # Vérifier l'avertissement d'anonymisation
        self.assertIn('anonymisées', export_data['avertissement'])
        
        # Vérifier que les statistiques sont présentes
        self.assertEqual(
            export_data['statistiques']['nombre_exploitations'],
            100
        )


class ReportGenerationServiceTests(TestCase):
    """
    Tests pour le service de génération de rapports
    Exigences: 25.5, 25.6
    """
    
    def setUp(self):
        """Configuration des données de test"""
        from .services import ReportGenerationService
        self.service = ReportGenerationService
        
        # Données de test
        self.statistics_data = {
            'statistiques_globales': {
                'nombre_exploitations': 150,
                'superficie_totale_hectares': 2500.50,
                'emplois_crees': {
                    'total': 200,
                    'agronomes': 50,
                    'ouvriers': 150
                },
                'transactions': {
                    'volume': 500,
                    'valeur_totale_fcfa': 50000000.00,
                    'commission_plateforme_fcfa': 2500000.00
                }
            },
            'statistiques_par_region': [
                {
                    'region': {'id': 1, 'nom': 'Maritime', 'code': 'MAR'},
                    'nombre_exploitations': 80,
                    'superficie_totale_hectares': 1200.00,
                    'emplois_crees': {'total': 100, 'agronomes': 25, 'ouvriers': 75},
                    'transactions': {'volume': 250, 'valeur_totale_fcfa': 25000000.00}
                }
            ],
            'repartition_transactions': [
                {
                    'type': 'ACHAT_DOCUMENT',
                    'nombre_transactions': 200,
                    'montant_total_fcfa': 1000000.00,
                    'commission_totale_fcfa': 0.00
                },
                {
                    'type': 'RECRUTEMENT_AGRONOME',
                    'nombre_transactions': 100,
                    'montant_total_fcfa': 10000000.00,
                    'commission_totale_fcfa': 1000000.00
                }
            ]
        }
    
    def test_generate_excel_report(self):
        """Test de la génération de rapport Excel"""
        excel_buffer = self.service.generate_excel_report(self.statistics_data)
        
        # Vérifier que le buffer n'est pas vide
        self.assertIsNotNone(excel_buffer)
        self.assertGreater(excel_buffer.getbuffer().nbytes, 0)
        
        # Vérifier que c'est un fichier Excel valide
        from openpyxl import load_workbook
        excel_buffer.seek(0)
        wb = load_workbook(excel_buffer)
        
        # Vérifier que le workbook a une feuille
        self.assertGreater(len(wb.sheetnames), 0)
        
        ws = wb.active
        
        # Vérifier le titre
        self.assertIn('RAPPORT', ws['A1'].value)
        
        # Vérifier l'avertissement d'anonymisation
        self.assertIn('anonymisées', ws['A3'].value.lower())
    
    def test_generate_pdf_report(self):
        """Test de la génération de rapport PDF"""
        pdf_buffer = self.service.generate_pdf_report(self.statistics_data)
        
        # Vérifier que le buffer n'est pas vide
        self.assertIsNotNone(pdf_buffer)
        self.assertGreater(pdf_buffer.getbuffer().nbytes, 0)
        
        # Vérifier que c'est un fichier PDF valide (commence par %PDF)
        pdf_buffer.seek(0)
        header = pdf_buffer.read(4)
        self.assertEqual(header, b'%PDF')


class ExportReportAPITests(TestCase):
    """
    Tests pour l'endpoint d'export de rapports
    Exigences: 25.5, 25.6
    """
    
    def setUp(self):
        """Configuration des données de test"""
        self.client = APIClient()
        
        # Créer une région
        self.region = Region.objects.create(nom="Maritime", code="MAR")
        self.prefecture = Prefecture.objects.create(
            nom="Golfe",
            code="GOL",
            region=self.region
        )
        self.canton = Canton.objects.create(
            nom="Lomé 1er",
            code="LOM1",
            prefecture=self.prefecture
        )
        
        # Créer un utilisateur institutionnel
        self.institution_user = User.objects.create_user(
            username='institution1',
            phone_number='+22890000010',
            user_type='INSTITUTION',
            password='Test1234!',
            two_factor_enabled=True
        )
        self.institution_profile = InstitutionProfile.objects.create(
            user=self.institution_user,
            nom_organisme='Ministère de l\'Agriculture',
            niveau_acces='NATIONAL'
        )
        
        # Créer des données de test
        exploitant_user = User.objects.create_user(
            username='exploitant1',
            phone_number='+22890000001',
            user_type='EXPLOITANT',
            password='Test1234!'
        )
        ExploitantProfile.objects.create(
            user=exploitant_user,
            superficie_totale=Decimal('15.50'),
            canton_principal=self.canton,
            coordonnees_gps={'lat': 6.1319, 'lon': 1.2228},
            statut_verification='VERIFIE'
        )
    
    def test_export_requires_authentication(self):
        """Test que l'export nécessite une authentification"""
        response = self.client.post(
            '/api/v1/institutional/reports/export/',
            {'format': 'excel'}
        )
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)
    
    def test_export_requires_institutional_user(self):
        """Test que l'export nécessite un compte institutionnel"""
        regular_user = User.objects.create_user(
            username='regular1',
            phone_number='+22890000011',
            user_type='EXPLOITANT',
            password='Test1234!'
        )
        self.client.force_authenticate(user=regular_user)
        
        response = self.client.post(
            '/api/v1/institutional/reports/export/',
            {'format': 'excel'}
        )
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
    
    def test_export_excel_success(self):
        """Test d'export Excel réussi"""
        self.client.force_authenticate(user=self.institution_user)
        
        response = self.client.post(
            '/api/v1/institutional/reports/export/',
            {'format': 'excel'}
        )
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        self.assertIn('attachment', response['Content-Disposition'])
        self.assertIn('.xlsx', response['Content-Disposition'])
    
    def test_export_pdf_success(self):
        """Test d'export PDF réussi"""
        self.client.force_authenticate(user=self.institution_user)
        
        response = self.client.post(
            '/api/v1/institutional/reports/export/',
            {'format': 'pdf'}
        )
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertIn('attachment', response['Content-Disposition'])
        self.assertIn('.pdf', response['Content-Disposition'])
    
    def test_export_invalid_format(self):
        """Test d'export avec format invalide"""
        self.client.force_authenticate(user=self.institution_user)
        
        response = self.client.post(
            '/api/v1/institutional/reports/export/',
            {'format': 'invalid'}
        )
        
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('error', response.data)
    
    def test_export_with_filters(self):
        """Test d'export avec filtres"""
        self.client.force_authenticate(user=self.institution_user)
        
        response = self.client.post(
            '/api/v1/institutional/reports/export/',
            {
                'format': 'excel',
                'region': self.region.id,
                'include_regions': True,
                'include_transactions': True
            }
        )
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
    
    def test_export_data_is_anonymized(self):
        """Test que les données exportées sont anonymisées"""
        self.client.force_authenticate(user=self.institution_user)
        
        response = self.client.post(
            '/api/v1/institutional/reports/export/',
            {'format': 'excel'}
        )
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        # Vérifier que le fichier Excel contient l'avertissement d'anonymisation
        from openpyxl import load_workbook
        from io import BytesIO
        
        excel_buffer = BytesIO(response.content)
        wb = load_workbook(excel_buffer)
        ws = wb.active
        
        # Chercher l'avertissement d'anonymisation
        found_warning = False
        for row in ws.iter_rows(max_row=10):
            for cell in row:
                if cell.value and 'anonymisées' in str(cell.value).lower():
                    found_warning = True
                    break
            if found_warning:
                break
        
        self.assertTrue(found_warning, "L'avertissement d'anonymisation n'a pas été trouvé dans le rapport")
